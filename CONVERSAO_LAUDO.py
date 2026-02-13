# %%
import sys
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import re
import unicodedata
import os

# %%
# ===================== PARÂMETROS =====================
PASTA_ENTRADAS = Path(
    r"C:\Users\migue\OneDrive - Agrorobotica Fotonica Em Certificacoes Agroambientais\_Fertilidade\ENTRADAS\ENTRADA_TESTE"
)
PASTA_SAIDAS = Path(
    r"C:\Users\migue\OneDrive - Agrorobotica Fotonica Em Certificacoes Agroambientais\_Fertilidade\ENTRADAS\SAIDA_TESTE\VERIFICAR"
)

ARQ_TEMPLATE = r"C:\Users\migue\OneDrive - Agrorobotica Fotonica Em Certificacoes Agroambientais\_Fertilidade\Template_Laudo.xlsx"
ARQ_ENDERECOS = r"C:\Users\migue\OneDrive - Agrorobotica Fotonica Em Certificacoes Agroambientais\_Fertilidade\DADOS_COMPLETO.xlsx"

NOME_ARQUIVO_SUST_MESTRE = r"C:\Users\migue\OneDrive - Agrorobotica Fotonica Em Certificacoes Agroambientais\_Fertilidade\ENTRADAS\ENTRADA_TESTE\F202255-SUST.xlsx"

# Abas do arquivo -UNICO (0-based)
SHEET_IDX_ABA1 = 1
SHEET_IDX_ABA2 = 2
SHEET_IDX_ABA3 = 3
HEADER_ROW_INDEX = 0  # linha 8

# Template
SHEET_INDEX_TEMPLATE = 0
HEADER_ROW_TEMPLATE = 1
DATA_START_ROW = 2

# Merge (lógica original): ABA1 base, ABA2/3 complementam por QR/PONTO/PROF
KEYS_MERGE = ["ID_AMOSTRA (QRCod)", "ID_PONTO", "PROFUNDIDADE"]

# ===================== FUNÇÕES DE NOME DE ARQUIVO =====================
def normalize_talhao_for_filename(value) -> str:
    if pd.isna(value):
        return "NA"
    try:
        if float(value).is_integer():
            value = int(float(value))
    except Exception:
        pass
    s = str(value).strip().replace(" ", "")
    s = re.sub(r"[^A-Za-z0-9_-]+", "", s)
    return s if s else "NA"


def build_output_name_from_input(base_filename_stem: str, talhao) -> str:
    stem_norm = base_filename_stem.replace("-UNICO", "").strip()
    talhao_str = normalize_talhao_for_filename(talhao)
    return f"{stem_norm}-T{talhao_str}.xlsx"


# ===================== NORMALIZAÇÃO FORTE DE CABEÇALHO =====================
_SUB_SUP_MAP = str.maketrans({
    "₀": "0", "₁": "1", "₂": "2", "₃": "3", "₄": "4", "₅": "5", "₆": "6", "₇": "7", "₈": "8", "₉": "9",
    "⁰": "0", "¹": "1", "²": "2", "³": "3", "⁴": "4", "⁵": "5", "⁶": "6", "⁷": "7", "⁸": "8", "⁹": "9",
    "¯": "", "−": "-", "–": "-", "—": "-",
})


def _strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))


def norm_header(x) -> str:
    if x is None:
        return ""
    s = str(x)
    s = s.translate(_SUB_SUP_MAP)
    s = _strip_accents(s).lower()
    s = re.sub(r"\s+", " ", s)         # \n \r \t
    s = re.sub(r"[^\w]+", " ", s)      # símbolos
    s = re.sub(r"\s+", " ", s).strip()
    return s


def build_col_lookup(df: pd.DataFrame) -> dict:
    lookup = {}
    for c in df.columns:
        k = norm_header(c)
        if k and k not in lookup:
            lookup[k] = c
    return lookup


def pick_col_flexible(lookup: dict, candidates: list[str]):
    # 1) match direto
    for cand in candidates:
        if isinstance(cand, str) and cand.startswith("re:"):
            continue
        key = norm_header(cand)
        if key in lookup:
            return lookup[key]
    # 2) regex em cabeçalho normalizado
    for cand in candidates:
        if isinstance(cand, str) and cand.startswith("re:"):
            rx = re.compile(cand[3:].strip())
            for nk in lookup.keys():
                if rx.search(nk):
                    return lookup[nk]
    return None


def read_sheet_full(xlsx_path: Path, sheet_index: int) -> pd.DataFrame:
    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_index, header=HEADER_ROW_INDEX, engine="openpyxl")

        if df is not None:
            print(f"Colunas lidas em {xlsx_path.name}: {df.columns.to_list()}")
            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
            df = df.dropna(how = "all").reset_index(drop = True)
            return df
        else:
            return pd.DataFrame()
        
    except Exception as e:
        print(f"[ERRO] Ao ler '{xlsx_path.name}' (aba índice {sheet_index}): {e}")

        return pd.DataFrame()



def map_columns_by_name(df_raw: pd.DataFrame, spec: dict, required: list[str], label: str) -> pd.DataFrame:
    lookup = build_col_lookup(df_raw)
    out = pd.DataFrame()
    missing = []
    for target, candidates in spec.items():
        real = pick_col_flexible(lookup, candidates)
        if real is None:
            out[target] = pd.NA
            if target in required:
                missing.append((target, candidates))
        else:
            out[target] = df_raw[real]

    if missing:
        print(f"[ERRO] ({label}) Não encontrei colunas obrigatórias. Detalhes:")
        for t, cands in missing:
            print(f"  - {t} (procurei por: {cands})")
        print("\n[INFO] Cabeçalhos detectados (normalizados):")
        for k in sorted(lookup.keys()):
            print("  -", k)
        sys.exit(1)

    return out.reset_index(drop=True)


# ===================== NORMALIZAÇÃO DAS CHAVES =====================
def normalize_ponto(v):
    if pd.isna(v):
        return pd.NA
    s = str(v).strip()
    m = re.search(r"\d+", s)
    if not m:
        return s
    return str(int(m.group(0)))


def normalize_profundidade(v):
    if pd.isna(v):
        return pd.NA
    s = str(v).strip()
    nums = re.findall(r"\d+", s)
    if len(nums) >= 2:
        a = int(nums[0])
        b = int(nums[1])
        return f"{a:02d}-{b:02d}"
    return s


def normalize_merge_keys(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "ID_PONTO" in df.columns:
        df["ID_PONTO"] = df["ID_PONTO"].apply(normalize_ponto)
    if "PROFUNDIDADE" in df.columns:
        df["PROFUNDIDADE"] = df["PROFUNDIDADE"].apply(normalize_profundidade)
    if "ID_AMOSTRA (QRCod)" in df.columns:
        df["ID_AMOSTRA (QRCod)"] = df["ID_AMOSTRA (QRCod)"].astype(str).str.strip()
        df.loc[df["ID_AMOSTRA (QRCod)"].isin(["nan", "None", ""]), "ID_AMOSTRA (QRCod)"] = pd.NA
    return df


# ===================== CORREÇÃO 171: PONTO VINDO COMO PROFUNDIDADE =====================
DEPTH_PAT = re.compile(r"^\s*\d{1,2}\s*-\s*\d{1,2}\s*$")


def is_depth_like(x) -> bool:
    if pd.isna(x):
        return False
    return bool(DEPTH_PAT.match(str(x).strip()))


def fix_ponto_when_depthlike(df1: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Se ID_PONTO estiver depth-like (ex.: 00-20) em vez de 1,2,3...,
    inferimos ponto por TALHAO_COMERCIAL, assumindo 3 profundidades por ponto,
    e retornamos df1 corrigido + mapa QR->PONTO.
    """
    df1 = df1.copy()
    if "PROFUNDIDADE" in df1.columns:
        df1["PROFUNDIDADE"] = df1["PROFUNDIDADE"].apply(normalize_profundidade)

    if "ID_PONTO" not in df1.columns or "TALHAO_COMERCIAL" not in df1.columns:
        return df1, {}

    mask_bad = df1["ID_PONTO"].apply(is_depth_like) & df1["PROFUNDIDADE"].apply(is_depth_like)
    if not mask_bad.any():
        return df1, {}

    qr_to_ponto = {}

    # Para cada talhão comercial, ordena e atribui ponto a cada bloco de 3 profundidades
    for talhao, g in df1.loc[mask_bad].groupby("TALHAO_COMERCIAL", dropna=False):
        g2 = g.copy()

        def _prof_key(v):
            if pd.isna(v):
                return 9999
            m = re.findall(r"\d+", str(v))
            return int(m[0]) if m else 9999

        g2["_prof_key"] = g2["PROFUNDIDADE"].apply(_prof_key)
        g2 = g2.sort_values(by=["ID_AMOSTRA (QRCod)", "_prof_key"], kind="mergesort")
        g2 = g2.drop(columns=["_prof_key"])

        ponto_inferido = (np.arange(len(g2)) // 3) + 1
        g2["_PONTO_INFERIDO"] = ponto_inferido.astype(int).astype(str)

        for qr, p in zip(g2["ID_AMOSTRA (QRCod)"], g2["_PONTO_INFERIDO"]):
            if pd.notna(qr):
                qr_to_ponto[str(qr).strip()] = p

        df1.loc[g2.index, "ID_PONTO"] = g2["_PONTO_INFERIDO"]

    return df1, qr_to_ponto


def apply_qr_point_map(df: pd.DataFrame, qr_to_ponto: dict) -> pd.DataFrame:
    """
    Aplica QR->PONTO em df2/df3 quando ID_PONTO vier depth-like.
    """
    if not qr_to_ponto or "ID_PONTO" not in df.columns:
        return df

    df = df.copy()
    df["PROFUNDIDADE"] = df["PROFUNDIDADE"].apply(normalize_profundidade)

    mask_bad = df["ID_PONTO"].apply(is_depth_like) & df["PROFUNDIDADE"].apply(is_depth_like)
    if not mask_bad.any():
        return df

    qrs = df.loc[mask_bad, "ID_AMOSTRA (QRCod)"].astype(str).str.strip()
    df.loc[mask_bad, "ID_PONTO"] = qrs.map(qr_to_ponto).fillna(df.loc[mask_bad, "ID_PONTO"])
    return df


# ===================== ORDENAÇÃO FINAL =====================
def _to_int_or_nan(x):
    try:
        if pd.isna(x):
            return np.nan
        m = re.search(r"\d+", str(x))
        return int(m.group(0)) if m else np.nan
    except Exception:
        return np.nan


def depth_order_key(depth):
    if pd.isna(depth):
        return 9999
    nums = re.findall(r"\d+", str(depth))
    if not nums:
        return 9999
    return int(nums[0])


def sort_dataframe_for_output(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ordena:
    1) TALHAO_COMERCIAL (alfabético)
    2) ID_PONTO (numérico)
    3) PROFUNDIDADE (00-20,20-40,40-60)
    """
    df = df.copy()

    # garante string p/ ordem alfabética
    if "TALHAO_COMERCIAL" in df.columns:
        df["_TALHAO_STR"] = df["TALHAO_COMERCIAL"].astype(str).str.strip().str.upper()
        df.loc[df["_TALHAO_STR"].isin(["NAN", "NONE", ""]), "_TALHAO_STR"] = "ZZZ"
    else:
        df["_TALHAO_STR"] = "ZZZ"

    df["_PONTO_NUM"] = df["ID_PONTO"].apply(_to_int_or_nan) if "ID_PONTO" in df.columns else np.nan
    df["_PROF_ORD"] = df["PROFUNDIDADE"].apply(depth_order_key) if "PROFUNDIDADE" in df.columns else 9999

    df = df.sort_values(by=["_TALHAO_STR", "_PONTO_NUM", "_PROF_ORD"], kind="mergesort").reset_index(drop=True)
    return df.drop(columns=["_TALHAO_STR", "_PONTO_NUM", "_PROF_ORD"], errors="ignore")


# ===================== CÁLCULOS =====================
def ensure_numeric(df: pd.DataFrame, cols: list[str]):
    for c in cols:
        if c in df.columns:
            df[c] = df[c].apply(lambda x: pd.to_numeric(x, errors = 'ignore') if pd.notna(x) else x)

def to_math(series):
    """Converte para numérico apenas para cálculo, tratando strings como zero."""
    return pd.to_numeric(series, errors='coerce').fillna(0)

def frac_pct(numerador: pd.Series, denominador: pd.Series):
    num = pd.to_numeric(numerador, errors="coerce").astype(float)
    den = pd.to_numeric(denominador, errors="coerce").astype(float)
    with np.errstate(divide="ignore", invalid="ignore"):
        return np.where(den > 0, (num / den) * 100.0, np.nan)


def safe_div(a: pd.Series, b: pd.Series):
    aa = pd.to_numeric(a, errors="coerce").astype(float)
    bb = pd.to_numeric(b, errors="coerce").astype(float)
    with np.errstate(divide="ignore", invalid="ignore"):
        return np.where(bb != 0, aa / bb, np.nan)


# ===================== SPECS (POR NOME) =====================
SPEC_ABA1 = {
    "ID_AMOSTRA (QRCod)": ["QR CODE", "QR-CODE"],
    "ID_TALHAO": ["Talhão Agrorobotica", "Talhao Agrorobotica", "Talhao Agro", "Área"],
    "TALHAO_COMERCIAL": ["Talhão Comercial", "Talhao Comercial"],
    "ID_PONTO": ["Ponto de Coleta", "Ponto Coleta", "# de Coleta"],
    "PROFUNDIDADE": ["Profundidade (cm)", "Profundidade cm", "Prof.\n(cm)", "Prof.(cm)", "Prof. (cm)"],

    "pH_CaCl2": ["pH CaCl2", "pH CaCl₂", "pH\nCaCl2", "pH\nCaCl₂"],
    "pH_agua": ["pH H2O", "pH H₂O", "pH\nH2O", "pH\nH₂O"],

    "P_(meh)_(mg/dm³)": [
        "P(meh) (mg dm-3)", "P(meh)\n(mg dm-3)",
        "P (mg dm-3)", "P\n(mg dm-3)",
        r"re:^p(?!\s+res)\b.*\bmg\b\s+dm\s+3$",
    ],
    "P_(res)_(mg/dm³)": [
        "P(res) (mg dm-3)", "P(res)\n(mg dm-3)", "P(res) \n(mg dm¯³)",
        r"re:^p\s+res\b.*\bmg\b\s+dm\s+3$",
    ],

    "S_(mg/dm³)": ["S (mg dm-3)", "S\n(mg dm-3)"],
    "MOS_(g/dm³)": ["MO (g dm-3)", "MO\n(g dm-3)"],

    "Ca_(cmolc/dm³)": ["Ca (cmolc dm-3)", "Ca\n(cmolc dm-3)"],
    "Mg_(cmolc/dm³)": ["Mg (cmolc dm-3)", "Mg\n(cmolc dm-3)"],
    "K_(cmolc/dm³)": ["K (cmolc dm-3)", "K\n(cmolc dm-3)"],
    "K_(mg/dm³)": ["K (mg dm-3)", "K\n(mg dm-3)"],

    "Al_(cmolc/dm³)": ["Al (cmolc dm-3)", "Al\n(cmolc dm-3)"],
    "H+Al_(cmolc/dm³)": ["H+Al (SMP)\n(cmolc dm-3)", "H+Al (SMP) (cmolc dm-3)", "H+Al (SMP)"],
}

SPEC_ABA2 = {
    "ID_AMOSTRA (QRCod)": ["QR CODE", "QR-CODE"],
    "ID_PONTO": ["Ponto de Coleta", "Ponto Coleta", "# de Coleta"],
    "PROFUNDIDADE": ["Profundidade (cm)", "Profundidade cm", "Prof.\n(cm)", "Prof.(cm)", "Prof. (cm)"],

    "B_(mg/dm³)": ["Boro\n(mg dm-3)", "Boro (mg dm-3)"],
    "Cu_(mg/dm³)": ["Cobre\n(mg dm-3)", "Cobre (mg dm-3)"],
    "Fe_(mg/dm³)": ["Ferro\n(mg dm-3)", "Ferro (mg dm-3)"],
    "Mn_(mg/dm³)": ["Manganês\n(mg dm-3)", "Manganês (mg dm-3)", "Manganes (mg dm-3)"],
    "Zn_(mg/dm³)": ["Zinco\n(mg dm-3)", "Zinco (mg dm-3)"],
}

SPEC_ABA3 = {
    "ID_AMOSTRA (QRCod)": ["QR CODE", "QR-CODE"],
    "ID_PONTO": ["Ponto de Coleta", "Ponto Coleta", "# de Coleta"],
    "PROFUNDIDADE": ["Profundidade (cm)", "Profundidade cm", "Prof.\n(cm)", "Prof.(cm)", "Prof. (cm)"],

    "Areia_(g/dm³)": ["Areia\n(g kg-1)", "Areia (g kg-1)", "Areia\n(g dm-3)", "Areia (g dm-3)", "Areia"],
    "Argila_(g/dm³)": ["Argila\n(g kg-1)", "Argila (g kg-1)", "Argila\n(g dm-3)", "Argila (g dm-3)"],
    "Silte_(g/dm³)": ["Silte\n(g kg-1)", "Silte (g kg-1)", "Silte\n(g dm-3)", "Silte (g dm-3)"],
}

REQUIRED_ABA1 = [
    "ID_AMOSTRA (QRCod)", "ID_TALHAO", "TALHAO_COMERCIAL", "ID_PONTO", "PROFUNDIDADE",
    "pH_CaCl2",
    "S_(mg/dm³)", "MOS_(g/dm³)",
    "Ca_(cmolc/dm³)", "Mg_(cmolc/dm³)", "K_(cmolc/dm³)",
    "Al_(cmolc/dm³)", "H+Al_(cmolc/dm³)",
]
REQUIRED_ABA2 = ["ID_AMOSTRA (QRCod)", "ID_PONTO", "PROFUNDIDADE"]
REQUIRED_ABA3 = ["ID_AMOSTRA (QRCod)", "ID_PONTO", "PROFUNDIDADE"]


# ===================== LEITURAS =====================
def read_aba1_quimica(caminho: Path) -> pd.DataFrame:
    df_raw = read_sheet_full(caminho, SHEET_IDX_ABA1)
    df = map_columns_by_name(df_raw, SPEC_ABA1, REQUIRED_ABA1, "ABA 1")

    if not df.empty:

        mask_sujeira = (
            df["ID_AMOSTRA (QRCod)"].astype(str).str.contains("Nota", case=False, na=False) | (df["ID_PONTO"].isna() & df["PROFUNDIDADE"].isna())
        )

        df = df[~mask_sujeira].reset_index(drop=True)

    df = normalize_merge_keys(df)

    ensure_numeric(df, [
        "pH_CaCl2", "pH_agua", "P_(meh)_(mg/dm³)", "P_(res)_(mg/dm³)",
        "S_(mg/dm³)", "MOS_(g/dm³)", "Ca_(cmolc/dm³)", "Mg_(cmolc/dm³)", 
        "K_(cmolc/dm³)", "K_(mg/dm³)", "Al_(cmolc/dm³)", "H+Al_(cmolc/dm³)",
    ])

    # P flexível: se não existir P(meh), usa P(res)
    has_meh = df["P_(meh)_(mg/dm³)"].notna().any()
    has_res = df["P_(res)_(mg/dm³)"].notna().any()
    if not (has_meh or has_res):
        print("[ERRO] (ABA 1) Não encontrei fósforo: nem P(meh)/P genérico, nem P(res).")
        sys.exit(1)
    if (not has_meh) and has_res:
        df["P_(meh)_(mg/dm³)"] = df["P_(res)_(mg/dm³)"]

    # Recálculo (mantém sua lógica)
    ca_v = to_math(df["Ca_(cmolc/dm³)"])
    mg_v = to_math(df["Mg_(cmolc/dm³)"])
    kk_v = to_math(df["K_(cmolc/dm³)"])
    al_v = to_math(df["Al_(cmolc/dm³)"])
    hal_v = to_math(df["H+Al_(cmolc/dm³)"])

    df["SB_(cmolc/dm³)"] = (ca_v + mg_v + kk_v).round(3)
    df["CTCp_(cmolc/dm³)"] = (df["SB_(cmolc/dm³)"] + hal_v).round(3)
    df["CTCe_(cmolc/dm³)"] = (df["SB_(cmolc/dm³)"] + al_v).round(3)

    # Nome que o template espera
    df["T_(cmolc/dm³)"] = df["CTCp_(cmolc/dm³)"]

    # Saturações (V% e m%)
    df["V_(%)"] = frac_pct(df["SB_(cmolc/dm³)"], df["CTCp_(cmolc/dm³)"]).round(2)
    df["Sat._Al_(%)"] = frac_pct(al_v, df["CTCe_(cmolc/dm³)"]).round(2)

    # Relações (usando valores numéricos para evitar erro de string)
    df["Ca/CTC_(%)"] = frac_pct(ca_v, df["CTCp_(cmolc/dm³)"]).round(2)
    df["Mg/CTC_(%)"] = frac_pct(mg_v, df["CTCp_(cmolc/dm³)"]).round(2)
    df["K/CTC_(%)"] = frac_pct(kk_v, df["CTCp_(cmolc/dm³)"].round(2))
    df["H+Al/CTC_(%)"] = frac_pct(hal_v, df["CTCp_(cmolc/dm³)"].round(2))

    df["Ca/Mg"] = safe_div(ca_v, mg_v).round(2)
    df["Ca/K"] = safe_div(ca_v, kk_v).round(2)
    df["Mg/K"] = safe_div(mg_v, kk_v).round(2)

    return df.reset_index(drop=True)


def read_aba2_micro(caminho: Path) -> pd.DataFrame:
    df_raw = read_sheet_full(caminho, SHEET_IDX_ABA2)
    df = map_columns_by_name(df_raw, SPEC_ABA2, REQUIRED_ABA2, "ABA 2")
    df = normalize_merge_keys(df)
    ensure_numeric(df, ["B_(mg/dm³)", "Cu_(mg/dm³)", "Fe_(mg/dm³)", "Mn_(mg/dm³)", "Zn_(mg/dm³)"])
    return df.reset_index(drop=True)


def read_aba3_fisica(caminho: Path) -> pd.DataFrame:
    df_raw = read_sheet_full(caminho, SHEET_IDX_ABA3)
    df = map_columns_by_name(df_raw, SPEC_ABA3, REQUIRED_ABA3, "ABA 3")
    df = normalize_merge_keys(df)
    ensure_numeric(df, ["Areia_(g/dm³)", "Argila_(g/dm³)", "Silte_(g/dm³)"])
    return df.reset_index(drop=True)


# ===================== MERGE (ABA1 BASE) =====================
def merge_left(base: pd.DataFrame, addon: pd.DataFrame, keys: list[str], label: str) -> pd.DataFrame:
    for k in keys:
        if k not in base.columns:
            print(f"[ERRO] {label}: chave '{k}' não existe no dataframe base.")
            sys.exit(1)
        if k not in addon.columns:
            print(f"[ERRO] {label}: chave '{k}' não existe no dataframe complementar.")
            sys.exit(1)

    out = pd.merge(base, addon, on=keys, how="left", suffixes=("", "_r"))
    dup = [c for c in out.columns if c.endswith("_r")]
    if dup:
        out = out.drop(columns=dup)
    return out


# ===================== SUST =====================
def carregar_mapa_sust(caminho_sust: Path) -> pd.DataFrame:
    try:
        df = pd.read_excel(caminho_sust, sheet_name=0, header=0, usecols="B,C,D,I", engine="openpyxl")
    except Exception as e:
        print(f"[ERRO] Ao ler SUST '{caminho_sust.name}': {e}")
        sys.exit(1)

    colnames = [c.strip() if isinstance(c, str) else c for c in df.columns]
    try:
        idx_coordx = colnames.index("Coord-X")
        idx_coordy = colnames.index("Coord-Y")
        idx_qrcode = colnames.index("QR-Code")
    except ValueError:
        print("[ERRO] O SUST precisa ter 'Coord-X', 'Coord-Y' e 'QR-Code' nas colunas B, C e I.")
        sys.exit(1)

    df_mapa = pd.DataFrame({
        "QR-Code": df.iloc[:, idx_qrcode],
        "Coord-X": df.iloc[:, idx_coordx],
        "Coord-Y": df.iloc[:, idx_coordy],
        "COL_D": df.iloc[:, 2],
    })
    return df_mapa.dropna(subset=["QR-Code"]).drop_duplicates(subset=["QR-Code"])


# ===================== ENDEREÇOS =====================
def carregar_endereco_por_id(caminho_enderecos: Path, id_codigo: str):
    try:
        df = pd.read_excel(caminho_enderecos, sheet_name=0, header=0, usecols="A,B", engine="openpyxl")
    except Exception as e:
        print(f"[ERRO] Ao ler ENDERECOS '{caminho_enderecos}': {e}")
        sys.exit(1)

    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
    if "ID" not in df.columns:
        print("[ERRO] ENDERECOS.XLSX precisa ter a coluna 'ID' na coluna A.")
        sys.exit(1)

    col_B_name = df.columns[1]
    df["ID"] = df["ID"].astype(str).str.strip()
    linha = df.loc[df["ID"] == str(id_codigo).strip()]

    if linha.empty:
        print(f"[AVISO] ID '{id_codigo}' não encontrado em ENDERECOS.XLSX. Coluna C ficará vazia.")
        return None

    val = linha.iloc[0][col_B_name]
    if pd.isna(val) or (isinstance(val, str) and val.strip() == ""):
        return None
    return val


# ===================== FORMATAÇÃO =====================
def _apply_fmt_range(ws, col_start_letter: str, col_end_letter: str, fmt: str, start_row: int, nrows: int):
    if nrows <= 0:
        return
    c1 = column_index_from_string(col_start_letter)
    c2 = column_index_from_string(col_end_letter)
    for r in range(start_row, start_row + nrows):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).number_format = fmt


def set_number_formats(ws, data_start_row: int, nrows: int):
    _apply_fmt_range(ws, "K", "AC", "0.00", data_start_row, nrows)
    _apply_fmt_range(ws, "AO", "AQ", "0.00", data_start_row, nrows)
    _apply_fmt_range(ws, "AH", "AN", "0", data_start_row, nrows)


# ===================== TEMPLATE HELPERS =====================
def header_map_from_template(ws, header_row: int) -> dict:
    mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col_idx).value
        name = (str(v).strip() if v is not None else "")
        if name:
            mapping[name] = col_idx
    return mapping


def copy_row_style(ws, src_row: int, dst_row: int):
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col_idx)
        dst = ws.cell(row=dst_row, column=col_idx)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
        col_letter = get_column_letter(col_idx)
        try:
            ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
        except Exception:
            pass


def ensure_rows_with_style(ws, needed_rows: int, header_row: int, data_start_row: int):
    current_capacity = ws.max_row - header_row
    if current_capacity >= needed_rows:
        return
    to_add = needed_rows - current_capacity
    style_ref_row = max(ws.max_row, data_start_row)
    insert_at = ws.max_row + 1
    ws.insert_rows(insert_at, amount=to_add)
    for r in range(insert_at, insert_at + to_add):
        copy_row_style(ws, style_ref_row, r)


# ===================== WRITE =====================
def write_group_to_template(df_sub: pd.DataFrame, caminho_tpl: Path, caminho_saida: Path,
                            df_sust_map: pd.DataFrame, nome_sust_sem_sufixo: str,
                            endereco_val):
    try:
        wb = load_workbook(caminho_tpl)
        ws = wb.worksheets[SHEET_INDEX_TEMPLATE]
    except Exception as e:
        print(f"[ERRO] Ao carregar template: {e}")
        sys.exit(1)

    # Ordena ANTES de escrever
    df_sub = sort_dataframe_for_output(df_sub)

    tpl_map = header_map_from_template(ws, HEADER_ROW_TEMPLATE)
    cols_to_write = [c for c in df_sub.columns if c in tpl_map]

    needed_rows = len(df_sub)
    if needed_rows > 0:
        ensure_rows_with_style(ws, needed_rows, HEADER_ROW_TEMPLATE, DATA_START_ROW)

    # Escrever dados do corpo
    for i, (_, row) in enumerate(df_sub.iterrows(), start=0):
        excel_row = DATA_START_ROW + i
        for col_name in cols_to_write:
            col_idx = tpl_map[col_name]
            val = row[col_name]
            ws.cell(row=excel_row, column=col_idx).value = None if (
                pd.isna(val) if not isinstance(val, str) else val == ""
            ) else val

    # A-F via SUST
    mapa = df_sust_map.set_index("QR-Code")
    for i, (_, row) in enumerate(df_sub.iterrows(), start=0):
        excel_row = DATA_START_ROW + i
        qr = row.get("ID_AMOSTRA (QRCod)")
        coord_x = coord_y = col_d = None

        if pd.notna(qr) and qr in mapa.index:
            reg = mapa.loc[qr]
            coord_x = reg["Coord-X"]
            coord_y = reg["Coord-Y"]
            col_d = reg["COL_D"]

        ws.cell(row=excel_row, column=1).value = None if pd.isna(coord_x) else coord_x
        ws.cell(row=excel_row, column=2).value = None if pd.isna(coord_y) else coord_y
        ws.cell(row=excel_row, column=3).value = None if pd.isna(endereco_val) else endereco_val
        ws.cell(row=excel_row, column=4).value = "Agrorobótica"
        ws.cell(row=excel_row, column=5).value = None if pd.isna(col_d) else col_d
        ws.cell(row=excel_row, column=6).value = nome_sust_sem_sufixo

    # Extras M/AP/AQ
    ca = pd.to_numeric(df_sub.get("Ca_(cmolc/dm³)"), errors="coerce")
    mg = pd.to_numeric(df_sub.get("Mg_(cmolc/dm³)"), errors="coerce")
    kk = pd.to_numeric(df_sub.get("K_(cmolc/dm³)"), errors="coerce")

    m_vals = (ca + mg).round(2)
    with np.errstate(divide="ignore", invalid="ignore"):
        ap_vals = np.where(kk > 0, (ca / kk), np.nan)
        aq_vals = np.where(kk > 0, (mg / kk), np.nan)
    ap_vals = np.round(ap_vals, 2)
    aq_vals = np.round(aq_vals, 2)

    col_M = column_index_from_string("M")
    col_AP = column_index_from_string("AP")
    col_AQ = column_index_from_string("AQ")

    for i in range(needed_rows):
        excel_row = DATA_START_ROW + i
        ws.cell(row=excel_row, column=col_M).value = None if pd.isna(m_vals.iloc[i]) else float(m_vals.iloc[i])
        ws.cell(row=excel_row, column=col_AP).value = None if pd.isna(ap_vals[i]) else float(ap_vals[i])
        ws.cell(row=excel_row, column=col_AQ).value = None if pd.isna(aq_vals[i]) else float(aq_vals[i])

    set_number_formats(ws, DATA_START_ROW, needed_rows)
    wb.save(caminho_saida)

def main():
    caminho_tpl = Path(ARQ_TEMPLATE)
    caminho_end = Path(ARQ_ENDERECOS)
    # NOME_ARQUIVO_SUST_MESTRE deve ser o caminho completo definido por você
    caminho_sust = Path(NOME_ARQUIVO_SUST_MESTRE) 

    if not caminho_sust.exists():
        print(f"[ERRO] SUST mestre não encontrado: {caminho_sust}")
        return

    # 1. Carrega o mapa SUST e informações globais
    print(f"[INFO] Carregando SUST mestre: {caminho_sust.name}")
    df_sust_map = carregar_mapa_sust(caminho_sust)
    
    # Define o nome do arquivo final com base no nome do SUST (sem o -SUST)
    nome_base_sust = caminho_sust.stem.replace("-SUST", "") 
    nome_master = f"{nome_base_sust}.xlsx"

    # 2. Localiza os arquivos -UNICO
    arquivos_unico = list(PASTA_ENTRADAS.glob("*-UNICO.xlsx"))
    arquivos_unico.sort() # Mantém a ordem alfabética/numérica
    
    if not arquivos_unico:
        print(f"[AVISO] Nenhum arquivo '-UNICO.xlsx' encontrado.")
        return

    print(f"[INFO] Concatenando {len(arquivos_unico)} arquivos para o Master: {nome_master}")
    PASTA_SAIDAS.mkdir(parents=True, exist_ok=True)

    lista_para_concatenar = []
    endereco_referencia = None

    for caminho_p1 in arquivos_unico:
        base_stem = caminho_p1.stem
        try:
            # Busca o endereço (ex: F20251861S)
            id_endereco = base_stem.replace("-UNICO", "")
            endereco_val = carregar_endereco_por_id(caminho_end, id_endereco)
            
            if endereco_referencia is None and endereco_val is not None:
                endereco_referencia = endereco_val

            # Leitura das abas com as limpezas de "Nota" na Aba 1
            df1 = read_aba1_quimica(caminho_p1)
            df2 = read_aba2_micro(caminho_p1)
            df3 = read_aba3_fisica(caminho_p1)

            # Correções de ponto/profundidade
            df1, qr_to_ponto = fix_ponto_when_depthlike(df1)
            df2 = apply_qr_point_map(df2, qr_to_ponto)
            df3 = apply_qr_point_map(df3, qr_to_ponto)

            # Merge das abas limpas
            df_all = merge_left(df1, df2, KEYS_MERGE, "MERGE 1+2")
            df_all = merge_left(df_all, df3, KEYS_MERGE, "MERGE Final")

            lista_para_concatenar.append(df_all)
            print(f" > Processado: {base_stem}")

        except Exception as e:
            print(f"[ERRO] Falha ao processar {base_stem}: {e}")

    # --- FINALIZAÇÃO: GERAÇÃO DO MASTER CONSOLIDADO ---
    if lista_para_concatenar:
        df_consolidado_final = pd.concat(lista_para_concatenar, ignore_index=True)
        
        # Ordenação final do bloco completo
        df_consolidado_final = sort_dataframe_for_output(df_consolidado_final)

        caminho_master = PASTA_SAIDAS / nome_master
        
        # Grava no Excel usando o Template para manter layout, estilos e fórmulas
        write_group_to_template(
            df_consolidado_final, 
            caminho_tpl, 
            caminho_master, 
            df_sust_map, 
            nome_base_sust, 
            endereco_referencia
        )
        print(f"\n[SUCESSO] Master gerado: {nome_master} com {len(df_consolidado_final)} linhas.")

    print("[CONCLUÍDO]")

if __name__ == "__main__":
    main()
